import re
import requests
from openpyxl import  Workbook
import time
import json
import random

class Douban_Spider:

    def __init__(self,addr):

        self.addr=addr
        self.url_pattern='https://www.douban.com/tag/{0}/book?start={1}'
        self.wb=Workbook()
        self.ws=[]

    # 爬虫主体，输入想爬取的标签和页数
    def Spider_body(self,tags,page):
        start_time=time.time()

        for i in range(len(tags)):
            print('正在爬取标签%s' % tags[i])
            self.ws.append(self.wb.create_sheet(title=tags[i]))
            for j in range(page):

                items=self.get_info(tags[i],j)  #爬取内容
                contents = self.format_info(items)  # 格式化内容
                self.write_excel(contents, i)   #将数据保存与excle 中
                # self.write_txt(contents)   #将数据保存在txt中

                time.sleep(random.randint(1,5))

        self.wb.save(self.addr)
        print('共耗时%.2f 秒' %(time.time()-start_time))


    def get_info(self,tag,page):
        hds = [
            {'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'}, \
            {
                'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'}, \
            {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]
        url=self.url_pattern.format(tag,page)
        try:
            r=requests.get(url,headers=hds[page %3])
            if r.status_code==200:
                print('请求成功')
            else:print('请求失败')
        except Exception as e:
            print('请求异常')

        pattern = re.compile(
            '<a.*?class="title.*?_blank">(.*?)</a>.*?class="desc">(.*?)</div>.*?rating_nums">(.*?)</span>', re.S)
        items=re.findall(pattern,r.text)

        if items:
            print("提取出页面内容，等待格式化与输出")
        else:
            print('正则表达式有问题')
        return items


    def format_info(self,items):
        contents=[]
        for item in items:
            contents.append((item[0],item[1].strip(),item[2]))
        return contents

    def write_excel(self,contents,i):
        for content in contents:
            self.ws[i].append(content)

    #
    # def write_txt(self,contents):
    #     with open(self.addr+'txt','a+',encoding='utf-8') as f:
    #         for content in contents:
    #             f.write(json.dumps(content,ensure_ascii=False)+'\n')


if __name__=='__main__':
    spider = Douban_Spider('E:/Desktop/try.xlsx') #输入你想保存的地址
    tags=['小说','爱情','历史'] #输入标签和页码后即可运行
    pages=80
    spider.Spider_body(tags, pages)









        
        



